import openpyxl

#creating a workbook
wb = openpyxl.load_workbook('target.xlsx')

#using the product sheet in excel
ws = wb.active

num = int(input('enter serial number\n'))

index = [ws.cell(row = 1, column = i).value for i in range(1,ws.max_column)]

values = [ws.cell(row = num+1 , column = i).value for i in range(1,ws.max_column)]




for val in range(0,ws.max_column):
    if val < ws.max_column:
        print(val)
        val+=1
        print(index[val] ," = ",values[val],"\n")
