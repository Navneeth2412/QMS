from flask import Flask, render_template, request ,redirect, url_for , send_file , flash, session
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook, Workbook
from datetime import date,datetime
from docxtpl import DocxTemplate
from docx import Document
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired, Length, EqualTo
from docx2pdf import convert
import os

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib 
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24).hex()

# Initialize ws as a global variable
ws = None

cust_detial = None

users = {'N0025': {'userid': 'N0025', 'password': 'ww',},
         'N0035':{'userid': 'N0035', 'password':'qq'}}




class LoginForm(FlaskForm):
    userid = StringField('User ID', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

class ResetPasswordForm(FlaskForm):
    userid = StringField('User ID', validators=[DataRequired()])
    submit = SubmitField('Reset Password')
userid = ''
@app.route('/', methods=['GET', 'POST'])
def login():
    global userid
    print("inside login")
    form = LoginForm()
    if form.validate_on_submit():
        print("submited login")
        userid = form.userid.data
        password = form.password.data
        print(f"Attempting login with UserID: {userid}")
        print("User ID:", userid)
        print("Password:", password)
        print("Users Dictionary:", users)
        # Check if user exists and password is correct (replace with database check)
        if userid in users and users[userid]['password'] == password:
            flash('', 'success')
            session['user'] = userid 
            return redirect(url_for('home'))
        else:
            flash('Invalid user id or password. Please try again.', 'danger')
    else:
        print(form.errors)
    return render_template('login.html', form=form)

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    # Clear the user session
    session.pop('user', None)
    flash('', 'success')
    # Redirect to the login page
    return redirect(url_for('login'))

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    form = ResetPasswordForm()
    if form.validate_on_submit():
        userid = form.userid.data

        # Check if user exists (replace with database check)
        if userid in users:
            # Generate a random password reset token (consider using a secure library)
            # Placeholder for generation, replace with a secure method (e.g., using secrets module)
            reset_token = os.urandom

            # Send a password reset email to the user's email address
            send_reset_email(users[userid]['email'], reset_token)

            flash('A password reset link has been sent to your email address.', 'success')
            return redirect(url_for('login'))
        else:
            flash('User not found. Please enter a valid user ID.', 'danger')

    return render_template('reset_password.html', form=form)



# def send_reset_email(email, reset_token):
#     # Replace with your actual email server configuration and ensure secure practices like TLS/SSL
#     # Consider using a secure library like Flask-Mail for email functionality

#     msg = MIMEMultipart()
#     msg['From'] = 'QMS Woodward'
#     msg['To'] = email
#     msg['Subject'] = 'Password Reset Request'

#     body = f'You have requested a password reset for your account. Please click the following link to reset your password:\n' \
#            f'{url_for("reset_password_with_token", reset_token=reset_token, _external=True)}'

#     msg.attach(MIMEText(body, 'plain'))

#     # Use a secure connection and port (e.g., 465 with SSL)
#     server = smtplib.SMTP_SSL('smtp.example.com', 465)
#     server.login('navneethpras@gmail.com', 'ugzk xyzd nhgx polo')
#     server.sendmail(msg['From'], msg['To'], msg.as_string())
#     server.quit()




#  ------------FUNCTIONS FOR COPYING AND DELETING----------------------
global cust_ref
global pno
global pname

def fill_template(data,offer,rev_name):
    
    template_path = 'template_quote.docx'
    template = DocxTemplate(template_path)

    # # Replace placeholders
    table_rows = []

    for data in data:
        # Create a dictionary for each row
        row_dict = {
            'slno': data['slno'],
            'part_no': data['part_no'],
            'description': data['description'],
            'qty': data['qty'],
            'amount': data['amount'],
            'total_price': data['total_price']
        }
        table_rows.append(row_dict)

    cust_row = []
    for offer in offer:
        cust_dict = {
            'custref':offer['custref'],
            'pno' : offer['pno'],
            'pname' : offer['pname'],
            'subject':offer['subject'],
            'inco':offer['inco'],
            'remark':offer['remark'],
            'total':offer['total'],
            'name':offer['name'],
            'phno':offer['phno'],
            'email':offer['email'],
            'cust_name' : offer['cust_name'],
            'cust_addr1' : offer['cust_addr1'],
            'cust_addr2' : offer['cust_addr2'],
            'cust_addr3' : offer['cust_addr3'],
            'cust_addr4' : offer['cust_addr4'],
            'city' : offer['city'],
            'cust_zip' : offer['cust_zip'],
            'country' : offer['country'],
            'contact_name' : offer['contact_name'],
            'designation' : offer['designation'],
            'cust_email' : offer['cust_email'],
            'cust_no' : offer['cust_no'],
            'cust_phno': offer['cust_phno'],

        }
    
    cust_row.append(cust_dict)
    print('this is cust_dict', cust_dict)
    print('this is cust_row', cust_row)


    print(table_rows)
    
    offerno = rev_name
    # Add the list of rows to the context
    context = {'table_rows': table_rows,
               'file' : offerno,
               'date' : d,
               'cust_row':cust_row,
               }
    

    
    # Render the template
    template.render(context)
    # Save the filled document
    output = rev_name
    output_path = output + '.docx'
    template.save(output_path)

    return output_path

@app.route('/generatepdf', methods=['POST'])
def generatepdf():
    if request.method=='POST':
        print("inside pdf")
        print(file)
        docx = file[:14] +'.docx'
        try:
            pdf_file = convert(docx)
            if pdf_file is not None:
                print("after convert")
                new_file = file[:14] + '.pdf'
                return send_file(pdf_file, as_attachment=True, mimetype='application/pdf', download_name=new_file)
            else:
                result_message = "Error generating PDF"
        except Exception as e:
            result_message = f"Error: {str(e)}"

        return render_template('index.html', filename=filename, date=d, date_time=d, result_message=result_message, sheet=ws)

    return render_template('index.html', filename=filename, date=d, date_time=d, result_message=None, sheet=ws)



@app.route('/generate', methods=['POST'])
def generate():
    print("filename:---------",filename)
    file = filename+'.xlsx'.strip()
    print("file:", file)
    wb = openpyxl.load_workbook(file)
    res = len(wb.sheetnames)
    if res >1:
        rev_name = file[:14] + '_' + 'R' + str(res-1)
    else:
        rev_name = file[:14]
    ws = wb.worksheets[res-1]


    saleswb = openpyxl.load_workbook('sales_emp.xlsx')
    salesws = saleswb.active

    custwb = openpyxl.load_workbook('QMS_Customer_Data_Sheet.xlsx',data_only=True)
    custws = custwb.active
    
    document = Document()
    if request.method == 'POST':
        off_list = []
        print("inside generate function")
        cust_ref = request.form['cust_ref']
        pno = request.form['p_no']
        pname = request.form['p_name']
        subject = request.form['subject']
        inco = request.form['inco']
        remark = request.form['remark']
        contact = request.form['contact']
        name =''
        phno = ''
        email = ''
        cust_name = ''
        cust_addr1 = ''
        cust_addr2 = ''
        cust_addr3 = ''
        cust_addr4 = ''
        city = ''
        cust_zip = ''
        country = ''
        contact_name = ''
        designation = ''
        cust_email = ''
        cust_no = ''
        cust_phno = ''

        print(custws.cell(row=4,column=1).value, type(custws.cell(row=4,column=1).value))
        print(userid, type(userid))
        for row in range(1, custws.max_row+1):
            if custws.cell(row=row,column=1).value == userid:
                print("inside customer loop")
                cust_name = custws.cell(row=row,column=3).value
                cust_addr1 = custws.cell(row=row,column=4).value
                cust_addr2 = custws.cell(row=row,column=5).value
                cust_addr3 = custws.cell(row=row,column=6).value
                cust_addr4 = custws.cell(row=row,column=7).value
                city = custws.cell(row=row,column=8).value
                cust_zip = custws.cell(row=row,column=9).value
                country = custws.cell(row=row,column=10).value
                contact_name = custws.cell(row=row,column=14).value
                designation = custws.cell(row=row,column=15).value
                cust_email = custws.cell(row=row,column=13).value
                cust_no = custws.cell(row=row,column=2).value
                cust_phno = custws.cell(row=row,column=12).value




        for row in range(1, salesws.max_row+1):

            if salesws.cell(row=row,column=1).value == contact:

                name = salesws.cell(row=row,column=1).value
                phno = salesws.cell(row=row,column=3).value
                email = salesws.cell(row=row,column=2).value
                break
       
        

       
        
        data_list = []
        total = 0
        for row in range(2, ws.max_row + 1):  # Assuming data starts from row 2
            total += float(ws.cell(row=row,column=6).value)
            data = {
                'slno': ws.cell(row=row, column=1).value,
                'part_no': ws.cell(row=row, column=2).value,
                'description': ws.cell(row=row, column=3).value,
                'qty': ws.cell(row=row, column=5).value,
                'amount': ws.cell(row=row, column=4).value,
                'total_price': ws.cell(row=row, column=6).value,
            }    
            document.add_paragraph()
            total = format(total,'.2f')
            data_list.append(data)
        
        print(data_list)
        print(total)

        offer = {
            'custref':cust_ref,
            'pno':pno,
            'pname':pname,
            'subject':subject,
            'inco':inco,
            'remark':remark,
            'total':total,
            'name':name,
            'phno':phno,
            'email':email,
            'cust_name' : cust_name,
            'cust_addr1' : cust_addr1,
            'cust_addr2' : cust_addr2,
            'cust_addr3' : cust_addr3,
            'cust_addr4' : cust_addr4,
            'city' : city,
            'cust_zip' : cust_zip,
            'country' : country,
            'contact_name' : contact_name,
            'designation' : designation,
            'cust_email' : cust_email,
            'cust_no' : cust_no,
            'cust_phno': cust_phno,


        }
        off_list.append(offer)
        # offer = []
        # offer.append(cust_ref)

        filled_doc_path = fill_template(data_list,off_list,rev_name)

        return send_file(filled_doc_path, as_attachment=True) 
    excel_files = get_excel_files()
    return render_template('index.html',filename=filename,date=d, date_time=d,result_message=None, sheet=ws,excel_files=excel_files)






def copy_row(part_data,target_sheet, quantity):

    # source_row = source_sheet[row_number]
    target_max = target_sheet.max_row+1
    target_column = target_sheet.max_column+1
    count = 1
    for cell in part_data:
        count+=1
        
        target_sheet.cell(row=target_max, column=count, value=cell)
    
    target_sheet.cell(row=target_max, column=target_column-2, value=quantity)
    print("printing this")
    print(target_sheet.cell(row=target_max, column=target_column-2).value)



def del_row(ws, row_nummber):
    # Your existing del_row function
    ws.delete_rows(row_nummber)

    return render_template('revise.html', result_message=None, sheet=ws)



#---------------------------------------CONVERTING TO WORDS----------------------------------------------------

def convert_to_words(num):  
    if num == 0:  
        return "Zero"  
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]  
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]  
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]  
    words = ""
    if num>= 1000:  
        words += ones[num // 1000] + " thousand "  
    num %= 1000  
    if num>= 100:  
        words += ones[num // 100] + " hundred "  
    num %= 100  
    if num>= 10 and num<= 19:  
        words += teens[num - 10] + " "  
    # num = 0  
    elif num>= 20:
        words += tens[num // 10] + " "  
    num %= 10  
    if num>= 1 and num<= 9:  
        words += ones[num] + " "  + "Only"
    return words.strip() 

#--------------------------------------FUNCTION FOR CREATING A NEW QUOTE-----------------------------------------
def get_next_letter(current_letter):
    if current_letter == 'Z':
        return 'A'
    else:
        return chr(ord(current_letter) + 1)

def find_next_available_letter(current_date, current_letter):
    while os.path.exists(f"{current_date}_{current_letter}_{userid}.xlsx"):
        current_letter = get_next_letter(current_letter)
    return current_letter







def create_file(user):
    current_date = date.today().strftime("%y%m%d")
    
    if not os.path.exists('last_state.txt'):
        with open('last_state.txt', 'w') as f:
            f.write(f"{current_date}_A")

        first_file = f"{current_date}_A_{user}.xlsx"

        wb = Workbook()
        wb.save(first_file)
        sheet1 = wb['Sheet']
        sheet1.title = 'R'
        wb.save(first_file)
    
        ws = wb.active
        headers = ['SERIAL NO', 'PART NUMBER', 'ITEM DESCRIPTION', 'UNIT PRICE', 'QTY', 'TOTAL PRICE']

        for cell, header in enumerate(headers, start=1):    
            ws.cell(row=1, column=cell, value=header)
        wb.save(first_file)
        return first_file
    
    with open('last_state.txt', 'r') as f:
        last_state = f.read().strip()

    last_date, last_letter = last_state.split('_')
    
    if last_date != current_date:
        next_letter = 'A'
    else:
        next_letter = find_next_available_letter(current_date, last_letter)

    new_state = f"{current_date}_{next_letter}"

    file = f"{new_state}_{user}.xlsx"

    with open('last_state.txt', 'w') as f:
        f.write(new_state)

    wb = Workbook()
    wb.save(file)

    sheet1 = wb['Sheet']
    sheet1.title = 'R'
    wb.save(file)
    
    ws = wb.active
    headers = ['SERIAL NO', 'PART NUMBER', 'ITEM DESCRIPTION', 'UNIT PRICE', 'QTY', 'TOTAL PRICE']

    for cell, header in enumerate(headers, start=1):    
        ws.cell(row=1, column=cell, value=header)
            
    wb.save(file)
    
    return file

# 


#-----------------------------ALL APP ROUTES---------------------------------

#------------------------------REVISE APP ROUTE------------------------------ ----------- ---------- ------------ --------- REVISE PAGE ROUTES ------------ ------------- ---------

@app.route('/revise/<file_name>', methods=['GET'])
def revise(file_name):
    global filename
    global filename1
    now = datetime.now()  # current date and time
    date_time = now.strftime("%I:%M %p")
    print("inside revise function")

    try:
        file = file_name
        wb = openpyxl.load_workbook(file)
        ws1 = wb.create_sheet()
        l = str(len(wb.sheetnames) - 1)
        ws1.title = 'R' + l
        filename = file[:14]
        work = wb.sheetnames[-2]
        ws = wb[work]

        for row in range(1, ws.max_row + 1):
            for cell in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=cell).value
                ws1.cell(row=row, column=cell, value=value)

        wb.save(file)

        res = len(wb.sheetnames)
        filename1 = filename + "_" + wb.sheetnames[-1]
        excel_files = get_excel_files()
        return render_template('revise.html', filename=filename1, date=d, date_time=date_time,result_message=None, sheet=ws1,excel_files=excel_files,cust_detail=cust_detail)
    except Exception as e:
        result_message = f"Error revising file: {str(e)}"
        return render_template('revise.html', filename=filename1, date=d, date_time=date_time,result_message=result_message, sheet=None,excel_files=excel_files,cust_detail=cust_detail)





# @app.route('/revise', methods=['POST'])
# def revise():
#     global file
#     global filename
#     global filename1
#     f = request.form['file']
#     file = f+".xlsx"
#     wb = openpyxl.load_workbook(file)
#     ws1 = wb.create_sheet()
#     l = str(len(wb.sheetnames)-1)
#     ws1.title = 'R'+l
#     filename = file[:14]
#     work = wb.sheetnames[-2]
#     ws = wb[work]
#     now = datetime.now() # current date and time
#     date_time = now.strftime("%I:%M %p")
#     print("inside revise function")
#     for row in range(1,ws.max_row+1):
#         for cell in range(1,ws.max_column+1):
#             value = ws.cell(row=row,column=cell).value
           
#             ws1.cell(row=row, column=cell,value=value)
#     wb.save(file)
  
#     res = len(wb.sheetnames)
    

#     filename1 = filename+"_" +wb.sheetnames[-1]

#     return render_template('revise.html',filename=filename1,date=d,date_time=date_time, result_message=None, sheet=ws1)

#------------------------------SHOW TABLE TO REVISE---------------------------

@app.route('/indexrev', methods=['POST'])
def indexrev():
    global ws  # Use the global ws variable
    if request.method == 'POST':
        wb = load_workbook(file)
        res = len(wb.sheetnames)
        print(res)
        ws = wb.worksheets[res-1]
        excel_files = get_excel_files()
    return render_template("revise.html",filename=filename,date=d, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)


#-------------------------------ADD PRODUCT FOR REVISE---------------------------

@app.route('/addrev', methods=['GET', 'POST'])
def addrev():
    global ws 
    global d
    wb = openpyxl.load_workbook(file)
    res = len(wb.sheetnames)
    ws = wb.worksheets[res-1]
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        part_no = request.form['part_no']
        quantity = request.form['quantity']


        # Load the source workbook
        source_workbook = load_workbook("price2.xlsx", read_only=True)
        source_sheet = source_workbook.active


        part_data = get_part_data(part_no)
        # Find the row with the specified Part ID and copy it to the target sheet

        
                # Copy the row and update quantity
        copy_row(part_data, ws, quantity)




        if ws.cell(row=ws.max_row-1,column=1).value == 'SERIAL NO':
            ws.cell(row=ws.max_row,column=1,value=1)
        else:
            count_row = int(ws.cell(row=ws.max_row-1,column=1).value)
            print("this is the serial number count",count_row)
            if count_row >= 1:
                    count = ws.cell(row=ws.max_row-1,column=1).value
                    ws.cell(row=ws.max_row,column=1,value=count+1)
            # else:
            #     ws.cell(row=ws.max_row,column=1,value=count)
                
        #         # Calculate total price and update in the target sheet
        price_column = 'D'
        result_column = 'F'
        quan_column = 'E'
                # Calculate total price and update in the target she

                # Ensure that the values are not None before converting to float

        quan_value = ws[quan_column + str(ws.max_row)].value
        print("this is quan value", quan_value)
        price = ws[price_column + str(ws.max_row)].value
        print(price)
        price2 = format(price,'.2f')
        ws[price_column + str(ws.max_row)].value = price2
        price_value = ws[price_column + str(ws.max_row)].value
        print("this is price value", price_value)

        if quan_value is not None and price_value is not None:
            total_pr = ws[result_column + str(ws.max_row)]
            total_pr = float(price_value) * float(quan_value)
            total_price = format(total_pr,'.2f')
            ws[result_column + str(ws.max_row)] = total_price
        else:
            # Handle the case where either quantity or price is None
            result_message = "Error: Quantity or Price is None."
            return render_template('index.html', result_message=result_message)       


            # Specify the output file path
        filename1 = file

            # Save the target workbook to a new file
        wb.save(filename1)
        result_message = f"Part details for ID {part_no} copied successfully."

        filename1 = filename+"_" +wb.sheetnames[-1]
        excel_files = get_excel_files()
        return render_template('revise.html',filename=filename1, result_message=None, sheet=ws,date_time=date_time, part_data = part_data,date= d,excel_files=excel_files,cust_detail=cust_detail)
    return render_template('revise.html',filename=filename1,  sheet=ws, part_data = part_data,date=d)

#------------------------------CANCEL FOR REVISION------------------------------

@app.route('/delrev', methods=['POST'])
def cancelrev():
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    wb = openpyxl.load_workbook(file)
    print(file)
    sheet = wb.sheetnames[-1]
    print(sheet)
    print(wb.sheetnames[-1])
    sheetname = wb.get_sheet_by_name(sheet)
    wb.remove_sheet(sheetname)
    wb.save(file)
    # if len(wb.sheetnames) <=2:
    #     filename = file[:14]] + "-" + wb.sheetnames[-1]
    # else:
    filename = file[:14] + "_" + wb.sheetnames[-1]
    excel_files = get_excel_files()
    return render_template('revise.html',filename=filename,date=d, date_time=date_time,result_message=None, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)

#-------------------------------UPDATE FOR REVISION------------------------------
@app.route('/update', methods=['POST'])
def update():
    wb = openpyxl.load_workbook(file)
    res = len(wb.sheetnames)
    ws = wb.worksheets[res-1]
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")

    slno = request.form['slno']
    print(slno)
    quant = request.form['quantity']

    price_column = 'D'

    price = float(ws[price_column + str(ws.max_row)].value)
    print(price)
    price2 = format(price,'.2f')
    ws[price_column + str(ws.max_row)].value = price2
    price_value = ws[price_column + str(ws.max_row)].value
    totalp = float(quant)*float(price_value)
    total_pr = format(totalp, '.2f')
    for cell in range(1,ws.max_row+1):
        if ws.cell(row=cell,column=1).value == int(slno):
            ws.cell(row=cell,column=5,value=quant)
            ws.cell(row=cell,column=6,value=total_pr)
    wb.save(file)
    excel_files = get_excel_files()
    return render_template("revise.html",filename=filename1,date=d, date_time=date_time,result_message=None, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)


#---------------------------------DELETE FOR REVISON------------------------------

@app.route('/deleterev' , methods=['POST'])
def deleterev():
    print("inside deleterev")
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")

    wb = openpyxl.load_workbook(file)
    res = len(wb.sheetnames)
    ws = wb.worksheets[res-1]
    # if request.method == ['POST']:
    part_no = request.form['slno']
    print("delete id ====", part_no, type(part_no))
        
        

    for row_number in range(2, ws.max_row   +1):
        print("row number ======" ,row_number)
        print(type(ws.cell(row=row_number, column=1).value))
        if ws.cell(row=row_number, column=1).value == int(part_no):
            del_row(ws, row_number)
            wb.save(file)
            break
    excel_files = get_excel_files()
    return render_template('revise.html', date=d,filename = filename1,date_time=date_time,result_message=None,sheet=ws,excel_files=excel_files,cust_detail=cust_detail)


#------------------------------TOTAL FOR REVISION-----------------------------------
@app.route('/totalrev', methods=['POST'])
def totalrev():
    global ws
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        wb = load_workbook(file)
        res = len(wb.sheetnames)
        ws = wb.worksheets[res-1]
        
        total = 0.00
        result_column = 'E'
        for cell in range(2,ws.max_row+1):
            t = float(ws.cell(row=cell,column=5).value)
            total += t

        ftotal = round(total,2)
        ws[result_column + str(ws.max_row+1)].value = ftotal

    for cell in range(1,ws.max_column):
        ws.cell(row=ws.max_row, column=cell, value=" ")
        
    ws.cell(row=ws.max_row,column=3, value="TOTAL PRICE")

    excel_files = get_excel_files()
    return render_template('revise.html',filename=filename1,date_time=date_time,date=d, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)



#--------------------------------ADD NEW QUOTATION----------------------------- ---------------------- ------------------ ------------------- NEW QUOATATION------------- ----------------- -----

@app.route('/createquote', methods=['POST'])
def createquote():
    global file
    print("inside creating quote function-------------------------------")
    file = create_file(userid)
    wb = load_workbook(file)
    res= len(wb.sheetnames)
    ws = wb.worksheets[res-1]
    filename = file[:14]
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    print("THIS IS FILE NAME============================",file)
    excel_files = get_excel_files()
    return render_template("add.html", result_message=None,date=d,date_time=date_time, filename= filename,sheet=ws,excel_files=excel_files,cust_detail=cust_detail)

#---------------------------------ADD NEW PRODUCT FOR QUOTE--------------------------

#----- FUNCTION FOR SEARCHING PART DESC -------

def get_part_data(part_id):
    wb = load_workbook('price2.xlsx')
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == part_id:
            return row
    return None
    

# ----- SEARCH ROUTE FOR PART DESC -------
@app.route('/search', methods=['POST'])
def addnew():
    global part_data
    print("in search")
    part_id = request.form.get('part_no')
    print(part_id)
    part_data = get_part_data(part_id)
    print(part_data)
    return {'part_data':part_data}




#--------------------------------DELETE APP ROUTE------------------------------  -------------- ----------- ------------- MAIN PAGE ROUTES -------- ------------- ----------- -----------
@app.route('/delete', methods=['POST'])
def delete():
    global ws  # Use the global ws variable
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        print('inside del func')
        del_id = request.form['del_id']
        print(del_id)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        print(file)
        print(ws.cell(row=2, column=1).value , type(ws.cell(row=2, column=1).value))
        print(type(del_id))
        for row_number in range(2, ws.max_row + 1):
            if ws.cell(row=row_number, column=2).value == del_id:
                del_row(ws, row_number)
                print('deleted')
                wb.save(file)
                break
        excel_files = get_excel_files()
        return render_template('index.html',filename=filename,date=d, date_time=date_time,result_message=None,sheet=ws,excel_files=excel_files,cust_detail=cust_detail)
    

#----------------------------VIEWING QUOTE-----------------------------------
    
@app.route('/view/<file_name>', methods=['GET'])
def view(file_name):
    global ws  # Use the global ws variable
    global filename
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    print("in view")
    print(file_name)
    
    try:
        file = file_name
        wb = load_workbook(file)
        res = len(wb.sheetnames)
        ws = wb.worksheets[res-1]
        if res == 1:
            filename1 = filename
        else:
            filename1 = filename + "-" + wb.sheetnames[-1]
    except Exception as e:
        result_message = "No File Found: " + str(e)
        return render_template("index.html", result_message=result_message, date_time=date_time, filename=filename, date=d, sheet=ws)
    excel_files = get_excel_files()
    return render_template("view.html", filename=filename1, date=d, date_time=date_time, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)



# @app.route('/view', methods=['POST'])
# def view():
#     global ws  # Use the global ws variable
#     global filename
#     now = datetime.now() # current date and time
#     date_time = now.strftime("%I:%M %p")
#     print("in view")
#     print(filename)
#     if request.method == 'POST':
#         res = 1
#         try:
#             f = request.form['file']
#             file = f+'.xlsx'
#             wb = load_workbook(file)
#             res = len(wb.sheetnames)
#             ws = wb.worksheets[res-1]
#             if res==1:
#                 filename1 = filename
#             else:
#                 filename1 = filename+"-" +wb.sheetnames[-1]
#         except:
#             result_message = "No File Found"
#             return render_template("index.html",result_message=result_message,date_time=date_time,filename=filename,date=d, sheet=ws)
#         print("this is file", file)

#     return render_template("view.html",filename=filename1,date=d,date_time=date_time, sheet=ws)


#-------------------------------ADDING NEW PRODUCT---------------------------------

#------ FUNCTION FOR SEARCHING PART DESC -------

def get_part_data(part_id):
    wb = load_workbook('price2.xlsx')
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == part_id:
            return row
    return None


#------------- SEARCH ROUTE ----------
@app.route('/search', methods=['POST'])
def search():
    global part_data
    print("in search")
    part_id = request.form.get('part_no')
    print(part_id)
    part_data = get_part_data(part_id)
    print(part_data)
    return {'part_data':part_data}


#------------------------------------ ROUTE TO DELETE LATEST QUOTE ------------------------------

@app.route('/deleted', methods=['POST'])

def cancel():
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    print(file)
    os.remove(file)
    excel_files = get_excel_files()
    return render_template('index.html',filename=filename,date_time=date_time,date=d, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)



#------------------------------------------ MAIN PAGE ROUTE -------------------------------------------------






@app.route('/home', methods=['GET', 'POST'])

def home():
    global ws 
    global file
    global filename
    global d
    global cust_ref
    global pno
    global pname
    global cust_detail

    user = session.get('user')

    # Check if a user is logged in
    if user:
        # Dummy dashboard route (replace with actual dashboard logic)

        cust  = openpyxl.load_workbook("QMS_Customer_Data_Sheet.xlsx",data_only=True)
        custws = cust.active

        file = create_file(user)
        print(file)
        cust_detail = []
        print(custws.cell(row=2,column=1).value,type(custws.cell(row=2,column=1).value))
        print(userid,type(userid))
        for row in range(1, custws.max_row+1):
            if custws.cell(row=row,column=1).value == userid:
                print("inside customer loop")
                for col in range(3,15):

                    cust_detail.append(custws.cell(row=row,column=col).value)
                break
        print(cust_detail)
        cust_detail = [item.strip() for item in cust_detail if item and isinstance(item, str) and item.strip()]

        print("this is customer detail", cust_detail)

        filename = file[:14]

        t = date.today()
                
        d= t.strftime("%d-%m-%Y")
        now = datetime.now() # current date and time
        date_time = now.strftime("%I:%M %p")
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        if request.method == 'POST':
            part_no = request.form['part_no']
            quantity = request.form['quantity']

            

            part_data = get_part_data(part_no)


            # Load the source workbook           mmmmv,v.                  
            source_workbook = load_workbook("price2.xlsx", read_only=True)
            source_sheet = source_workbook.active

            return render_template('index.html', result_message=None, date=d,filename=filename, sheet=ws,date_time=date_time, part_data=part_data)
    else:
        # Redirect to login if no user is logged in
        return redirect(url_for('login'))
    excel_files = get_excel_files()
    return render_template('index.html',filename=filename,date=d,date_time=date_time, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)



def get_excel_files():
    current_folder = os.path.dirname(os.path.abspath(__file__))
    excel_files = [file for file in os.listdir(current_folder) if file.endswith('.xlsx') and userid in file]
    return excel_files

#---------------------COPY DATA------------------------------------------------------------------


@app.route('/copy', methods=['GET', 'POST'])
def copy():
    global ws 
    global file
    global filename   
    global d

    print("in copy")
    wb = openpyxl.load_workbook(file)
    ws = wb.active 
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        part_no = request.form['part_no']
        quantity = request.form['quantity']


        # Load the source workbook
        source_workbook = load_workbook("price2.xlsx", read_only=True)
        source_sheet = source_workbook.active


        found_Part = False

        part_data = get_part_data(part_no)

        print(part_data)
        
      
        copy_row(part_data,ws,quantity)

        if ws.cell(row=ws.max_row-1,column=1).value == 'SERIAL NO':
            ws.cell(row=ws.max_row,column=1,value=1)
        else:
            count_row = int(ws.cell(row=ws.max_row-1,column=1).value)
            print("this is the serial number count",count_row)
            if count_row >= 1:
                    count = ws.cell(row=ws.max_row-1,column=1).value
                    ws.cell(row=ws.max_row,column=1,value=count+1)
            # else:
            #     ws.cell(row=ws.max_row,column=1,value=count)
                
        #         # Calculate total price and update in the target sheet
        price_column = 'D'
        result_column = 'F'
        quan_column = 'E'

        #         # Ensure that the values are not None before converting to float

        quan_value = ws[quan_column + str(ws.max_row)].value
        print("this is quan value", quan_value)
        price = ws[price_column + str(ws.max_row)].value
        print(price)
        price2 = format(price,'.2f')
        ws[price_column + str(ws.max_row)].value = price2
        price_value = ws[price_column + str(ws.max_row)].value
        print("this is price value", price_value)
        
        
        

        if quan_value is not None and price_value is not None:
            ws[result_column + str(ws.max_row)] = float(price_value) * float(quan_value)
            res = ws[result_column + str(ws.max_row)].value
            print(res)
            res2 = format(res, '.2f')
            print(res2)
            ws[result_column + str(ws.max_row)] = res2
        else:
                    # Handle the case where either quantity or price is None
            result_message = "Error: Quantity or Price is None."
            return render_template('index.html',filename=filename,date=d, result_message=result_message)
        found_Part = True

        if found_Part:
            # Specify the output file path
            filename1 = file

            # Save the target workbook to a new file
            wb.save(filename1)
            result_message = f"Part details for ID {part_no} copied successfully."
        else:
            result_message = f"Part details for ID {part_no} not found."
        filename = file[:14]
        excel_files = get_excel_files()
        return render_template('index.html', result_message=None,date_time=date_time,filename=filename,date=d, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)
    return render_template('index.html',filename=filename,date=d, sheet=ws)

#----------------------------------GETTING TOTAL PRICE-------------------------------

@app.route('/total', methods=['POST'])
def total():
    global ws

    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        wb = load_workbook(file)
        ws = wb.active
        total = 0.00
        result_column = 'E'
        for cell in range(2,ws.max_row+1):
            t = float(ws.cell(row=cell,column=5).value)
            total += t
            
       
        ftotal = round(total,2)
        ws[result_column + str(ws.max_row+1)].value = ftotal
    

    words = convert_to_words(int(total))
    print(words)
    

    for cell in range(1,ws.max_column):
        ws.cell(row=ws.max_row, column=cell, value=" ")
    
    ws.cell(row=ws.max_row,column=3,value=words)
    ws.cell(row=ws.max_row,column=2, value="TOTAL PRICE")
    excel_files = get_excel_files()
    
    return render_template('index.html',filename=filename,date=d,date_time=date_time, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)

#------------------------------SHOWS TABLE-----------------------------------------

@app.route('/index', methods=['POST'])
def index():
    global ws  # Use the global ws variable
    now = datetime.now() # current date and time
    date_time = now.strftime("%I:%M %p")
    if request.method == 'POST':
        wb = load_workbook(file)
        res = len(wb.sheetnames)
        ws = wb.worksheets[res-1]
    excel_files = get_excel_files()
    return render_template("index.html",filename=filename,date_time=date_time,date=d, sheet=ws,excel_files=excel_files,cust_detail=cust_detail)


#----------MAIN----------
if __name__ == '__main__':

    app.run(debug=True)

