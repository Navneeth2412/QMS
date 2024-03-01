# from flask import Flask, render_template, request, send_file
# from docx import Document
# from docxtpl import DocxTemplate

# app = Flask(__name__)

# def detect_table(doc):
#     tables = doc.tables
#     if tables:
#         # You can print information about each table or perform specific operations
#         for i, table in enumerate(tables):
#             print(f"Table {i + 1}: {len(table.rows)} rows x {len(table.columns)} columns")
#         return tables
#     else:
#         print("No tables found in the document")
#         return None

# def add_row_to_table(table, data):
#     new_row = table.add_row()
#     for cell, value in zip(new_row.cells, data):
#         cell.text = str(value)

# def fill_template(data):
#     template_path = 'template.docx'
    
#     # Open the template with python-docx to detect existing tables
#     doc = Document(template_path)
#     tables = detect_table(doc)

#     if tables:
#         # Assume you want to add a new row to the first table
#         detected_table = tables[0]

#         # Add a new row to the detected table
#         new_row_data = [4, 'Geek 4']  # Customize with your data
#         add_row_to_table(detected_table, new_row_data)

#         # Save the document with python-docx
#         doc.save('temp.docx')

#         # Reopen the template with docxtpl
#         template = DocxTemplate('temp.docx')
#     else:
#         # If no tables are found, use the original template with docxtpl
#         template = DocxTemplate(template_path)

#     # Replace placeholders
#     context = {key: str(value) for key, value in data.items()}

#     # Add table data to context
#     table_data = [
#         {'id': 1, 'name': 'Geek 1'},
#         {'id': 2, 'name': 'Geek 2'},
#         {'id': 3, 'name': 'Geek 3'}
#     ]
#     context['table_data'] = table_data

#     # Render the template using docxtpl
#     template.render(context)

#     # Save the filled document
#     output_path = 'filled_template.docx'
#     template.save(output_path)

#     return output_path

# @app.route('/')
# def index():
#     return render_template('template.html', filled_template=None)

# @app.route('/generate', methods=['POST'])
# def generate():
#     data = {
#         'Name': request.form.get('name'),
#         'Date': request.form.get('date'),
#         'Amount': request.form.get('amount')
#     }

#     filled_doc_path = fill_template(data)

#     return send_file(filled_doc_path, as_attachment=True)

# if __name__ == '__main__':
#     app.run(debug=True)



# Import docx NOT python-docx 
# import docx 
# import openpyxl
# # Create an instance of a word document 
# doc = docx.Document() 

# # Add a Title to the document 
# doc.add_heading('GeeksForGeeks', 0) 

# wb = openpyxl.load_workbook('240229_A_admin.xlsx')
# ws = wb.active

# # Table data in a form of list 
# data = 	((1, ws.cell(row=ws.max_row,column=3).value, 'hello'), 
# 	(2, 'Geek 2','he'), 
# 	(3, 'Geek 3','he'), )


# # d = ()

# # for cell in range(1,ws.max_row+1):
# #     for cols in range(1,6):
# #         string = str(ws.cell(row=cell,column=cols).value)
# #         d += (string)
# #     data += d
# # print(data)
# # Creating a table object 
# table = doc.add_table(rows=1, cols=6) 

# # Adding heading in the 1st row of the table 
# row = table.rows[0].cells 
# row[0].text = 'Id'
# row[1].text = 'Name'
# row[2].text = 'text'
# # Adding data from the list to the table 
# for id, name,text in data: 

# 	# Adding a row and then adding data in it. 
# 	row = table.add_row().cells 
# 	# Converting id to string as table can only take string input 
# 	row[0].text = str(id) 
# 	row[1].text = name 
# 	row[2].text = text
# table.style = 'Colorful List'
# # Now save the document to a location 
# doc.save('gfg.docx') 

#--------------------------------REAL ONE

# import docx 
# import openpyxl

# # Create an instance of a Word document 
# doc = docx.Document('template_quote.docx') 

# # Add a Title to the document 
# # doc.add_heading('Quote', 0) 

# # Load the Excel workbook and get the active sheet
# wb = openpyxl.load_workbook('240229_A_admin.xlsx')
# ws = wb.active

# # Table data in the form of a tuple of tuples
# data = []

# # Read all rows from Excel and add to the data list
# for row in ws.iter_rows(min_row=2, values_only=True):
#     data.append(row)

# # Creating a table object with dynamic columns based on the number of columns in Excel
# table = doc.add_table(rows=1, cols=len(data[0]))

# # Adding heading in the 1st row of the table 
# row = table.rows[0].cells 
# for idx, heading in enumerate(('S No.', 'Part No.', 'Description', 'Qty.','Price','Total Price')):
#     row[idx].text = heading

# # Adding data from the list to the table 
# for row_data in data: 
#     # Adding a row and then adding data in it.
#     row = table.add_row().cells 
#     # Converting data to string as the table can only take string input 
#     for idx, cell_data in enumerate(row_data):
#         row[idx].text = str(cell_data)

# # Apply a table style
# for row in table.rows:
#     for cell in row.cells:
#         cell.paragraphs[0].style = doc.styles['Table Grid']

# # table.style = 'Colorful List'
# for style in doc.styles:
#     print(style.name)

# # Now save the document to a location 
# doc.save('temp.docx') 





import docxtpl
import openpyxl

# Load the Excel workbook and get the active sheet
wb = openpyxl.load_workbook('240229_A_admin.xlsx')
ws = wb.active

# Create a list to hold data from Excel
data = []

# Read all rows from Excel and add to the data list
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)

# Assuming headers are present in the first row of the Excel file
headers = [col_header for col_header in ws[1]]

# Create a dictionary with data for the template
template_data = {
    'headers': headers,
    'table_data': data,
    'other_variable': 'some_value',  # Additional variables if needed
}

# Load your Word document template
doc = docxtpl.DocxTemplate('template_quote.docx')

# Render the template with the data
doc.render(template_data)

# Save the filled document
doc.save('document.docx')




