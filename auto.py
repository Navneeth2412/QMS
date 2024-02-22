from flask import Flask, request, render_template ,jsonify

import openpyxl



app = Flask(__name__)

@app.route('/')
def auto():
    wb = openpyxl.load_workbook('price.xlsx')
    ws = wb.active
    l = []
    for cell in range(2,ws.max_row+1):
        s=  ws.cell(row=cell, column=1).value
        l.append(s)

    return render_template("index.html", languages=l)

if __name__ == '__main__':
    app.run(debug=True)
